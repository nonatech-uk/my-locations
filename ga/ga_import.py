#!/usr/bin/env python3
"""
Import GA Flying Logbook from Excel to database.

Parses the 'Log Book' sheet from the Excel file and imports flight records.
"""

import sys
from pathlib import Path
from datetime import datetime, time, timedelta

# Add parent directory for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
import db

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def time_to_hours(t):
    """Convert datetime.time, timedelta, or numeric to decimal hours."""
    if t is None:
        return None
    if isinstance(t, (int, float)):
        return float(t)
    if isinstance(t, timedelta):
        return t.total_seconds() / 3600
    if isinstance(t, time):
        return t.hour + t.minute / 60 + t.second / 3600
    return None


def format_registration(reg):
    """Format registration with hyphen (GLSMI -> G-LSMI)."""
    if not reg:
        return None
    reg = reg.strip().upper()
    # UK registrations start with G, insert hyphen after first letter
    if reg and len(reg) >= 2 and reg[0] == 'G' and reg[1] != '-':
        return f"G-{reg[1:]}"
    return reg


def parse_logbook(xlsx_path):
    """
    Parse GA logbook Excel file.

    Returns list of flight record dictionaries.
    """
    print(f"Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Log Book']

    flights = []
    skipped = 0

    # Data rows start at 6, headers at row 5
    for row_idx in range(6, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=1).value

        # Skip rows without a date
        if not date_val:
            continue

        # Skip non-date values (sometimes there are summary rows)
        if not isinstance(date_val, datetime):
            skipped += 1
            continue

        # Extract all fields
        flight = {
            'date': date_val.date(),
            'aircraft_type': ws.cell(row=row_idx, column=5).value,
            'registration': format_registration(ws.cell(row=row_idx, column=6).value),
            'captain': ws.cell(row=row_idx, column=7).value,
            'operating_capacity': ws.cell(row=row_idx, column=8).value,
            'dep_airport': ws.cell(row=row_idx, column=9).value,
            'arr_airport': ws.cell(row=row_idx, column=10).value,
            'dep_time': ws.cell(row=row_idx, column=11).value,
            'arr_time': ws.cell(row=row_idx, column=12).value,
            # Hour categories - Single Engine
            'hours_sep_pic': time_to_hours(ws.cell(row=row_idx, column=13).value),
            'hours_sep_dual': time_to_hours(ws.cell(row=row_idx, column=14).value),
            # Hour categories - Multi Engine
            'hours_mep_pic': time_to_hours(ws.cell(row=row_idx, column=15).value),
            'hours_mep_dual': time_to_hours(ws.cell(row=row_idx, column=16).value),
            # Additional categories
            'hours_pic_3': time_to_hours(ws.cell(row=row_idx, column=17).value),
            'hours_dual_3': time_to_hours(ws.cell(row=row_idx, column=18).value),
            'hours_pic_4': time_to_hours(ws.cell(row=row_idx, column=19).value),
            'hours_dual_4': time_to_hours(ws.cell(row=row_idx, column=20).value),
            # Other fields
            # Column 21 may have instrument hours as decimal or time
            'hours_instrument': time_to_hours(ws.cell(row=row_idx, column=21).value),
            # Column 22: Hours as instructor (giving instruction)
            'hours_as_instructor': time_to_hours(ws.cell(row=row_idx, column=22).value),
            # Column 23: Simulator hours
            'hours_simulator': time_to_hours(ws.cell(row=row_idx, column=23).value),
            # Captain column contains instructor name for training flights
            'instructor': ws.cell(row=row_idx, column=7).value,
            # Column 25: Comments/exercise refs
            'exercise': ws.cell(row=row_idx, column=25).value,
            'hours_total': time_to_hours(ws.cell(row=row_idx, column=24).value),
        }

        # Clean up string fields
        for key in ['aircraft_type', 'captain', 'operating_capacity', 'dep_airport',
                    'arr_airport', 'instructor', 'exercise']:
            if flight[key] and isinstance(flight[key], str):
                flight[key] = flight[key].strip()

        flights.append(flight)

    print(f"Parsed {len(flights)} flights, skipped {skipped} rows")
    return flights


def import_ga_flights(flights, dry_run=False):
    """
    Import GA flight records to database.

    Args:
        flights: List of flight record dictionaries
        dry_run: If True, don't insert, just print what would be inserted
    """
    if dry_run:
        print("\nDry run - first 10 flights:")
        for f in flights[:10]:
            total = f['hours_total'] or 0
            print(f"  {f['date']} {f['aircraft_type']:5} {f['registration']:7} "
                  f"{f['dep_airport']}->{f['arr_airport']} {total:.2f}h "
                  f"({f['operating_capacity']}) {f['captain'] or ''}")
        print(f"\n... and {len(flights) - 10} more flights")
        return

    conn = db.get_connection()
    cur = conn.cursor()

    sql = """
        INSERT INTO ga_flights (
            date, aircraft_type, registration, captain, operating_capacity,
            dep_airport, arr_airport, dep_time, arr_time,
            hours_sep_pic, hours_sep_dual, hours_mep_pic, hours_mep_dual,
            hours_pic_3, hours_dual_3, hours_pic_4, hours_dual_4,
            hours_instrument, hours_as_instructor, hours_simulator,
            hours_total, instructor, exercise
        ) VALUES (
            %(date)s, %(aircraft_type)s, %(registration)s, %(captain)s, %(operating_capacity)s,
            %(dep_airport)s, %(arr_airport)s, %(dep_time)s, %(arr_time)s,
            %(hours_sep_pic)s, %(hours_sep_dual)s, %(hours_mep_pic)s, %(hours_mep_dual)s,
            %(hours_pic_3)s, %(hours_dual_3)s, %(hours_pic_4)s, %(hours_dual_4)s,
            %(hours_instrument)s, %(hours_as_instructor)s, %(hours_simulator)s,
            %(hours_total)s, %(instructor)s, %(exercise)s
        )
        ON CONFLICT (date, registration, dep_airport, arr_airport, dep_time)
        DO UPDATE SET
            aircraft_type = EXCLUDED.aircraft_type,
            captain = EXCLUDED.captain,
            operating_capacity = EXCLUDED.operating_capacity,
            hours_sep_pic = EXCLUDED.hours_sep_pic,
            hours_sep_dual = EXCLUDED.hours_sep_dual,
            hours_mep_pic = EXCLUDED.hours_mep_pic,
            hours_mep_dual = EXCLUDED.hours_mep_dual,
            hours_pic_3 = EXCLUDED.hours_pic_3,
            hours_dual_3 = EXCLUDED.hours_dual_3,
            hours_pic_4 = EXCLUDED.hours_pic_4,
            hours_dual_4 = EXCLUDED.hours_dual_4,
            hours_instrument = EXCLUDED.hours_instrument,
            hours_as_instructor = EXCLUDED.hours_as_instructor,
            hours_simulator = EXCLUDED.hours_simulator,
            hours_total = EXCLUDED.hours_total,
            instructor = EXCLUDED.instructor,
            exercise = EXCLUDED.exercise
    """

    inserted = 0
    errors = 0

    for flight in flights:
        try:
            cur.execute(sql, flight)
            inserted += 1
        except Exception as e:
            print(f"Error inserting {flight['date']} {flight['registration']}: {e}")
            errors += 1
            conn.rollback()
            continue

    conn.commit()
    cur.close()
    conn.close()

    print(f"\nInserted/updated {inserted} flights, {errors} errors")


def find_xlsx_file():
    """Find the logbook Excel file."""
    # Check data/ga directory first
    data_dir = Path(__file__).parent.parent / "data" / "ga"
    xlsx_files = list(data_dir.glob("*.xlsx"))
    if xlsx_files:
        return xlsx_files[0]

    # Check GA directory (original location)
    ga_dir = Path(__file__).parent.parent / "GA"
    xlsx_files = list(ga_dir.glob("*.xlsx"))
    if xlsx_files:
        return xlsx_files[0]

    return None


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Import GA flying logbook from Excel')
    parser.add_argument('xlsx_file', nargs='?', help='Path to Excel file')
    parser.add_argument('--dry-run', action='store_true', help='Parse but do not insert')
    args = parser.parse_args()

    if args.xlsx_file:
        xlsx_path = Path(args.xlsx_file)
    else:
        xlsx_path = find_xlsx_file()
        if not xlsx_path:
            print("No Excel file found. Specify path as argument.")
            sys.exit(1)

    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Importing from: {xlsx_path}")
    flights = parse_logbook(xlsx_path)

    if not flights:
        print("No flights parsed!")
        sys.exit(1)

    import_ga_flights(flights, dry_run=args.dry_run)


if __name__ == '__main__':
    main()
